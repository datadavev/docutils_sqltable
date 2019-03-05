"""
Created on Sep 7, 2011

@author: vieglais

Import a Excel worksheet as tables in an SQLite database.
"""

import logging
import sqlite3
import openpyxl


class Xls2Sql(object):
    """Utility class that will import worksheets from an Excel workbook as
    tables in an SQL database.
    """

    def __init__(self, db=None):
        """
        Initialize the class.

        :param db: An open DB-API 2.0 database connection. If None then an
                   in-memory SQLite3 database will be made.
        """
        self.log = logging.getLogger("Xls2Sql")
        self.db = db
        if self.db is None:
            self.db = sqlite3.connect(":memory:")

    def createTables(self, meta):
        """Create a new table in the database using the column names contained
        in meta.

        meta is structured like::

        meta = {name: {'start_row':
                     'columns': [{'name': column name,
                                  'id': zero offset from left,
                                  'type': sqlite type},
                                 ...],
                    },
            }
        """
        for tname in list(meta.keys()):
            sql = "CREATE TABLE %s (" % tname
            fields = []
            for col in meta[tname]["columns"]:
                fields.append("%s %s" % (col["name"], col["type"]))
            sql += ", ".join(fields)
            sql += ");"
            logging.info("Create table SQL: %s" % sql)
            cn = self.db.cursor()
            cn.execute(sql)
            self.db.commit()

    def getBookMeta(self, workbook, meta_row=0):
        """Load a meta structure from the workbook.

        :param workbook: An openpyxl workbook object.

        :param meta_row: Which row to read metadata (column names) from.
        """
        meta = {}
        for sname in workbook.sheetnames:
            sheet = workbook[sname]
            meta[sname] = self.getSheetMeta(sheet, meta_row)
        return meta

    def guessColumnType(self, row=1, col=0):
        """Place holder for code that will guess the type of the column. This is
        not really necessary for SQLite databases, as it figures stuff out on
        the fly.
        """
        return "text"

    def coerceType(self, val, typ):
        """Try and coerce the supplied value to the indicated type.

        :param val: The value to coerce

        :param typ: The target type
        """
        if typ == "text":
            return "'%s'" % str(val)
        elif typ == "real":
            return float(val)
        return str(val)

    def getSheetMeta(self, worksheet, meta_row=0):
        """Attempt to read column names for sheet from the specified row.

        :param worksheet: Worksheet being examined

        :param meta_row: Row of worksheet to be looked at.
        """
        meta = {"start_row": meta_row + 1, "columns": []}
        header = worksheet[meta_row + 1]
        for i in range(0, len(header)):
            cmeta = {}
            cmeta["name"] = str(header[i].value).lower()
            cmeta["id"] = i
            cmeta["type"] = self.guessColumnType(row=1, col=i)
            meta["columns"].append(cmeta)
        return meta

    def load(self, workbook, meta=None):
        """Load the supplied workbook into the database that was provided with
        the constructor.

        :param workbook: Path to an Excel file or an openpyxl workbook object

        :param meta: Optional metadata structure to use instead of auto generated.
        """
        if isinstance(workbook, str):
            workbook = openpyxl.reader.excel.load_workbook(workbook)
        if meta is None:
            meta = self.getBookMeta(workbook)
        self.createTables(meta)
        cn = self.db.cursor()
        for sname in list(meta.keys()):
            insert_string = []
            for col in meta[sname]["columns"]:
                insert_string.append("?")
            insert_string = ",".join(insert_string)
            sql = "INSERT INTO %s VALUES (%s);" % (sname, insert_string)
            sheet = workbook[sname]
            for i in range(meta[sname]["start_row"], sheet.max_row):
                row = sheet[i]
                rdata = []
                for col in meta[sname]["columns"]:
                    v = self.coerceType(row[col["id"]].value, col["type"])
                    rdata.append(v)
                self.log.debug(sql)
                cn.execute(sql, rdata)
                self.db.commit()
        return self.db


if __name__ == "__main__":
    pass
